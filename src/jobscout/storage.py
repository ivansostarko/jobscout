"""Excel storage with strict deduplication (one row per job, ever)."""
from __future__ import annotations

import logging
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Job

log = logging.getLogger("jobscout.storage")

COLUMNS = [
    "job_id", "scraped_at", "source", "title", "company", "location",
    "workplace_type", "category", "salary", "posted_at", "score", "url", "description",
]
WIDTHS = [18, 20, 12, 40, 28, 26, 12, 22, 16, 16, 8, 60, 80]


class ExcelStore:
    def __init__(self, path: str):
        self.path = Path(path)
        if not self.path.exists():
            self._create()

    def _create(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"
        ws.append(COLUMNS)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for i, _ in enumerate(COLUMNS, start=1):
            c = ws.cell(row=1, column=i)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
        ws.freeze_panes = "A2"
        wb.save(self.path)
        log.info("Created new Excel store at %s", self.path)

    def _load(self):
        return load_workbook(self.path)

    def existing_ids(self) -> set[str]:
        wb = self._load()
        ws = wb["Jobs"]
        return {str(r[0].value) for r in ws.iter_rows(min_row=2) if r[0].value}

    def add_jobs(self, jobs: list[Job]) -> list[Job]:
        """Append only jobs not already present. Returns the newly added jobs."""
        wb = self._load()
        ws = wb["Jobs"]
        seen = {str(r[0].value) for r in ws.iter_rows(min_row=2) if r[0].value}
        added: list[Job] = []
        for job in jobs:
            if job.job_id in seen:
                continue
            seen.add(job.job_id)
            d = job.to_dict()
            ws.append([d.get(col, "") for col in COLUMNS])
            added.append(job)
        if added:
            wb.save(self.path)
        log.info("Stored %d new jobs (of %d scraped)", len(added), len(jobs))
        return added

    def jobs_since(self, since: datetime) -> list[dict]:
        wb = self._load()
        ws = wb["Jobs"]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(COLUMNS, row))
            try:
                ts = datetime.fromisoformat(str(rec["scraped_at"]))
            except (ValueError, TypeError):
                continue
            if ts >= since:
                out.append(rec)
        return out

    def jobs_today(self) -> list[dict]:
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        return self.jobs_since(today)

    def jobs_this_week(self) -> list[dict]:
        now = datetime.now()
        monday = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        return self.jobs_since(monday)
